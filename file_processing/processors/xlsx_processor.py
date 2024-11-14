from io import BytesIO
from openpyxl import load_workbook
import msoffcrypto
from file_processing.errors import FileProcessingFailedError, FileCorruptionError
from file_processing.file_processor_strategy import FileProcessorStrategy

class XlsxFileProcessor(FileProcessorStrategy):
    """
    Processor for handling Excel (.xlsx) files, extracting metadata such as active sheet,
    sheet names, data, last modified by, and creator.

    Attributes:
        metadata (dict): Contains metadata fields including 'active_sheet', 'sheet_names', 'data',
                         'last_modified_by', 'creator', and 'has_password'.
    """

    def __init__(self, file_path: str, open_file: bool = True) -> None:
        """
        Initializes the XlsxFileProcessor with the specified file path.

        Args:
            file_path (str): Path to the .xlsx file to process.
            open_file (bool): Indicates whether to open and process the file immediately.

        Sets:
            metadata (dict): Populated with a message if `open_file` is False, otherwise initialized with default values.
        """
        super().__init__(file_path, open_file)
        self.metadata = {'message': 'File was not opened'} if not open_file else self._default_metadata()

    def _default_metadata(self) -> dict:
        """
        Returns default metadata for an unopened .xlsx file.

        Returns:
            dict: Default metadata fields including 'active_sheet', 'sheet_names', 'data',
                  'last_modified_by', 'creator', and 'has_password'.
        """
        return {
            "active_sheet": None,
            "sheet_names": None,
            "data": None,
            "last_modified_by": None,
            "creator": None,
            "has_password": False
        }

    def process(self) -> None:
        """
        Extracts metadata from the .xlsx file, including active sheet, sheet names, data,
        last modified by, creator, and checks for password protection.

        Raises:
            FileCorruptionError: If the file is corrupted or encrypted and cannot be opened.
            FileProcessingFailedError: If an error occurs during Excel file processing.
        """
        if not self.open_file:
            return

        with open(self.file_path, 'rb') as f:
            file_content = BytesIO(f.read())

        try:
            office_file = msoffcrypto.OfficeFile(file_content)
            if office_file.is_encrypted():
                self.metadata["has_password"] = True
                return
        except Exception as e:
            raise FileCorruptionError(f"File is corrupted: {self.file_path}") from e

        try:
            file_content.seek(0)  # Reset the position to the start
            exceldoc = load_workbook(self.file_path)
            self.metadata.update({
                "active_sheet": exceldoc.active.title,
                "sheet_names": exceldoc.sheetnames,
                "data": self.read_all_data(exceldoc),
                "last_modified_by": exceldoc.properties.lastModifiedBy,
                "creator": exceldoc.properties.creator
            })
        except Exception as e:
            raise FileProcessingFailedError(
                f"Error encountered while processing {self.file_path}: {e}"
            )

    def save(self, output_path: str = None) -> None:
        """
        Saves the .xlsx file with updated metadata to the specified output path.

        Args:
            output_path (str): Path to save the .xlsx file. If None, overwrites the original file.

        Raises:
            FileProcessingFailedError: If an error occurs while saving the .xlsx file.
        """
        try:
            exceldoc = load_workbook(self.file_path)
            cp = exceldoc.properties
            cp.creator = self.metadata.get('creator', cp.creator)
            cp.last_modified_by = self.metadata.get(
                'last_modified_by', cp.lastModifiedBy
            )

            save_path = output_path or self.file_path
            exceldoc.save(save_path)
        except Exception as e:
            raise FileProcessingFailedError(
                f"Error encountered while saving {self.file_path}: {e}"
            )

    @staticmethod
    def read_all_data(exceldoc):
        """
        Reads all data from each sheet in the .xlsx file and returns it as a dictionary.

        Args:
            exceldoc (Workbook): An instance of the Excel workbook.

        Returns:
            dict: Dictionary with sheet names as keys and lists of rows as values.

        Raises:
            FileProcessingFailedError: If an error occurs while reading data from the Excel document.
        """
        try:
            data = {}
            for sheet_name in exceldoc.sheetnames:
                sheet = exceldoc[sheet_name]
                sheet_data = []
                for row in sheet.iter_rows(values_only=True):
                    sheet_data.append(row)
                data[sheet_name] = sheet_data
            return data
        except Exception as e:
            raise FileProcessingFailedError(
                f"Error encountered while reading data from Excel document: {e}"
            )
