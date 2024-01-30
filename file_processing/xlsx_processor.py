from io import BytesIO
from openpyxl import load_workbook
import msoffcrypto
from file_processing.errors import FileProcessingFailedError, FileCorruptionError
from file_processing.file_processor_strategy import FileProcessorStrategy


class XlsxFileProcessor(FileProcessorStrategy):
    def __init__(self, file_path: str, open_file: bool = True) -> None:
        super().__init__(file_path, open_file)
        self.metadata = {
            'message': 'File was not opened'} if not open_file else self._default_metadata()

    def _default_metadata(self) -> dict:
        return {
            "active_sheet": None,
            "sheet_names": None,
            "data": None,
            "last_modified_by": None,
            "creator": None,
            "has_password": False
        }

    def process(self) -> None:
        if not self.open_file:
            return

        with open(self.file_path, 'rb') as f:
            file_content = BytesIO(f.read())

        try:
            office_file = msoffcrypto.OfficeFile(file_content)
            if office_file.is_encrypted():
                self.metadata["has_password"] = True
                return
        except Exception as e:
            raise FileCorruptionError(f"File is corrupted: {self.file_path}") from e

        try:
            file_content.seek(0)  # Reset the position to the start
            exceldoc = load_workbook(self.file_path)
            self.metadata.update({"active_sheet": exceldoc.active.title})
            self.metadata.update({"sheet_names": exceldoc.sheetnames})
            self.metadata.update({"data": self.read_all_data(exceldoc)})
            self.metadata.update(
                {"last_modified_by": exceldoc.properties.lastModifiedBy})
            self.metadata.update({"creator": exceldoc.properties.creator})
        except Exception as e:
            raise FileProcessingFailedError(
                f"Error encountered while processing {self.file_path}: {e}")

    def save(self, output_path: str = None) -> None:
        try:
            exceldoc = load_workbook(self.file_path)
            cp = exceldoc.properties
            cp.creator = self.metadata.get('creator', cp.creator)
            cp.last_modified_by = self.metadata.get(
                'last_modified_by', cp.lastModifiedBy)

            save_path = output_path or self.file_path
            exceldoc.save(save_path)
        except Exception as e:
            raise FileProcessingFailedError(
                f"Error encountered while saving {self.file_path}: {e}")

    @staticmethod
    def read_all_data(exceldoc):
        try:
            data = {}
            for sheet_name in exceldoc.sheetnames:
                sheet = exceldoc[sheet_name]
                sheet_data = []
                for row in sheet.iter_rows(values_only=True):
                    sheet_data.append(row)
                data[sheet_name] = sheet_data
            return data
        except Exception as e:
            raise FileProcessingFailedError(
                f"Error encountered while reading data from Excel document: {e}")
